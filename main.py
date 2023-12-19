from openpyxl import load_workbook

wb = load_workbook('bmi.xlsx')

sheet = wb.active
def kalkulator_bmi(weight, height):
    return weight / (height * height)


for row in sheet.iter_rows(min_row=2, min_col=1, max_row=21, max_col=3):
    name, weight, height = row[0].value, row[1].value, row[2].value
    bmi = kalkulator_bmi(weight, height)
    z_bmi = round(bmi,2)                                    #zaokraglone bmi do 2 miejsca po przecinku
    print(f"Osoba:{name}, BMI:{z_bmi}")

    if z_bmi < 16.00 :
        print("""Kategoria: wyglodzenie 
Waga ciala: niedowaga 
Ryzyko chorob: minimalne, ale zwiększony poziom wystąpienia innych problemów zdrowotnych""")

    if 16.99 >= z_bmi >= 16.00 :
        print("""Kategoria: wychudzenie 
Waga ciala: niedowaga 
Ryzyko chorob: minimalne, ale zwiększony poziom wystąpienia innych problemów zdrowotnych""")

    if 18.48 >= z_bmi >= 17.00 :
        print("""Kategoria: niedowaga 
Waga ciala: niedowaga 
Ryzyko chorob: minimalne, ale zwiekszony poziom wystąpienia innych problemów zdrowotnych""")

    if 24.99 >= z_bmi >= 18.50:
        print("""Kategoria :pozadana masa ciala
Waga ciala: optimum
Ryzyko chorob: minimalne """)

    if 29.99 >= z_bmi >= 25.00:
        print("""Kategoria :nadwaga
Waga ciala: nadwaga
Ryzyko chorob: srednie""")

    if 34.99 >= z_bmi >= 30.00:
        print("""Kategoria : otylosc I stopnia
Waga ciala: otylosc
Ryzyko chorob: wysokie """)

    if 39.99 >= z_bmi >= 35.00:
        print("""Kategoria : otylosc II stopnia
Waga ciala: otylosc
Ryzyko chorob: bardzo wysokie """)

    if z_bmi >= 40.00:
        print("""Kategoria : otylosc III stopnia
Waga ciala: otylosc
Ryzyko chorob: ekstremalny poziom ryzyka """)
